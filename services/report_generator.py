import os
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib import colors
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.lineplots import LinePlot
from reportlab.graphics.charts.barcharts import VerticalBarChart
from io import BytesIO
import base64

class ReportGenerator:
    """Gerador de relatórios em Excel e PDF"""
    
    def __init__(self):
        self.styles = getSampleStyleSheet()
        self.setup_custom_styles()
    
    def setup_custom_styles(self):
        """Configurar estilos personalizados para PDF"""
        self.styles.add(ParagraphStyle(
            name='CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            textColor=colors.HexColor('#0d6efd'),
            alignment=1  # Center
        ))
        
        self.styles.add(ParagraphStyle(
            name='SectionHeader',
            parent=self.styles['Heading2'],
            fontSize=16,
            spaceAfter=12,
            textColor=colors.HexColor('#198754'),
            borderWidth=1,
            borderColor=colors.HexColor('#198754'),
            borderPadding=5
        ))
    
    def generate_client_report_excel(self, clients, filename=None):
        """Gerar relatório de clientes em Excel"""
        if not filename:
            filename = f"relatorio_clientes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório de Clientes"
        
        # Cabeçalho
        headers = ['ID', 'Nome', 'Email', 'Telefone', 'CPF/CNPJ', 'Tipo de Seguro', 'Endereço', 'Data Cadastro', 'Status']
        ws.append(headers)
        
        # Estilizar cabeçalho
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Dados dos clientes
        for client in clients:
            ws.append([
                client.id,
                client.name,
                client.email,
                client.phone,
                client.cpf_cnpj,
                client.insurance_type or '-',
                client.address or '-',
                client.created_at.strftime('%d/%m/%Y'),
                client.status
            ])
        
        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Adicionar bordas
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=len(clients) + 1, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border
        
        # Salvar arquivo
        filepath = f"reports/{filename}"
        os.makedirs("reports", exist_ok=True)
        wb.save(filepath)
        return filepath
    
    def generate_sales_report_excel(self, cards, filename=None):
        """Gerar relatório de vendas em Excel"""
        if not filename:
            filename = f"relatorio_vendas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        wb = Workbook()
        
        # Aba principal - Resumo
        ws_summary = wb.active
        ws_summary.title = "Resumo Vendas"
        
        # Calcular estatísticas
        total_cards = len(cards)
        pipeline_stats = {}
        columns = ['atendimento_inicial', 'proposta_enviada', 'venda_andamento', 'venda_concluida', 'pos_venda']
        
        for column in columns:
            pipeline_stats[column] = len([c for c in cards if c.column == column])
        
        # Criar resumo
        summary_data = [
            ['Métrica', 'Valor'],
            ['Total de Operações', total_cards],
            ['Atendimento Inicial', pipeline_stats['atendimento_inicial']],
            ['Propostas Enviadas', pipeline_stats['proposta_enviada']],
            ['Vendas em Andamento', pipeline_stats['venda_andamento']],
            ['Vendas Concluídas', pipeline_stats['venda_concluida']],
            ['Pós-venda', pipeline_stats['pos_venda']],
            ['Taxa de Conversão (%)', round((pipeline_stats['venda_concluida'] / max(total_cards, 1)) * 100, 2)]
        ]
        
        for row in summary_data:
            ws_summary.append(row)
        
        # Estilizar resumo
        self._style_excel_sheet(ws_summary, summary_data)
        
        # Aba detalhada - Todas as operações
        ws_details = wb.create_sheet("Detalhes Operações")
        detail_headers = ['ID', 'Título', 'Descrição', 'Cliente ID', 'Responsável', 'Coluna', 'Prioridade', 'Criado em', 'Atualizado em']
        ws_details.append(detail_headers)
        
        for card in cards:
            ws_details.append([
                card.id,
                card.title,
                card.description,
                card.client_id or '-',
                card.assigned_to or '-',
                card.column.replace('_', ' ').title(),
                card.priority,
                card.created_at.strftime('%d/%m/%Y %H:%M'),
                card.updated_at.strftime('%d/%m/%Y %H:%M')
            ])
        
        self._style_excel_sheet(ws_details, None, detail_headers)
        
        # Adicionar gráficos
        self._add_excel_charts(wb, pipeline_stats)
        
        # Salvar arquivo
        filepath = f"reports/{filename}"
        os.makedirs("reports", exist_ok=True)
        wb.save(filepath)
        return filepath
    
    def _style_excel_sheet(self, worksheet, data=None, headers=None):
        """Aplicar estilos ao worksheet Excel"""
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
        
        # Estilizar primeira linha (cabeçalho)
        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Ajustar largura das colunas
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _add_excel_charts(self, workbook, pipeline_stats):
        """Adicionar gráficos ao Excel"""
        ws_charts = workbook.create_sheet("Gráficos")
        
        # Dados para gráfico
        chart_data = [
            ['Etapa', 'Quantidade'],
            ['Atendimento Inicial', pipeline_stats['atendimento_inicial']],
            ['Propostas Enviadas', pipeline_stats['proposta_enviada']],
            ['Vendas em Andamento', pipeline_stats['venda_andamento']],
            ['Vendas Concluídas', pipeline_stats['venda_concluida']],
            ['Pós-venda', pipeline_stats['pos_venda']]
        ]
        
        # Inserir dados
        for row in chart_data:
            ws_charts.append(row)
        
        # Criar gráfico de barras
        chart = BarChart()
        chart.title = "Pipeline de Vendas"
        chart.x_axis.title = "Etapas"
        chart.y_axis.title = "Quantidade"
        
        data = Reference(ws_charts, min_col=2, min_row=1, max_row=6, max_col=2)
        categories = Reference(ws_charts, min_col=1, min_row=2, max_row=6)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        
        ws_charts.add_chart(chart, "D2")
    
    def generate_client_report_pdf(self, clients, filename=None):
        """Gerar relatório de clientes em PDF"""
        if not filename:
            filename = f"relatorio_clientes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        filepath = f"reports/{filename}"
        os.makedirs("reports", exist_ok=True)
        
        doc = SimpleDocTemplate(filepath, pagesize=A4)
        story = []
        
        # Título
        title = Paragraph("Relatório de Clientes - Monteiro Corretora", self.styles['CustomTitle'])
        story.append(title)
        story.append(Spacer(1, 20))
        
        # Data do relatório
        date_p = Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}", self.styles['Normal'])
        story.append(date_p)
        story.append(Spacer(1, 20))
        
        # Resumo
        summary_title = Paragraph("Resumo Executivo", self.styles['SectionHeader'])
        story.append(summary_title)
        
        total_clients = len(clients)
        types_count = {}
        for client in clients:
            if client.insurance_type:
                types_count[client.insurance_type] = types_count.get(client.insurance_type, 0) + 1
        
        summary_text = f"""
        Total de clientes cadastrados: {total_clients}<br/>
        Tipos de seguro mais procurados: {', '.join([f"{k}: {v}" for k, v in sorted(types_count.items(), key=lambda x: x[1], reverse=True)[:3]])}<br/>
        Data de cadastro mais recente: {max([c.created_at for c in clients]).strftime('%d/%m/%Y') if clients else 'N/A'}
        """
        
        story.append(Paragraph(summary_text, self.styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Tabela de clientes
        clients_title = Paragraph("Lista Detalhada de Clientes", self.styles['SectionHeader'])
        story.append(clients_title)
        
        # Preparar dados da tabela
        table_data = [['Nome', 'Email', 'Telefone', 'Tipo Seguro', 'Data Cadastro']]
        
        for client in clients:
            table_data.append([
                client.name[:25] + '...' if len(client.name) > 25 else client.name,
                client.email[:30] + '...' if len(client.email) > 30 else client.email,
                client.phone,
                client.insurance_type or '-',
                client.created_at.strftime('%d/%m/%Y')
            ])
        
        # Criar tabela
        table = Table(table_data, colWidths=[2*inch, 2.5*inch, 1.5*inch, 1.5*inch, 1*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d6efd')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        
        story.append(table)
        
        # Construir PDF
        doc.build(story)
        return filepath
    
    def generate_sales_report_pdf(self, cards, monthly_data=None, filename=None):
        """Gerar relatório de vendas em PDF"""
        if not filename:
            filename = f"relatorio_vendas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        filepath = f"reports/{filename}"
        os.makedirs("reports", exist_ok=True)
        
        doc = SimpleDocTemplate(filepath, pagesize=A4)
        story = []
        
        # Título
        title = Paragraph("Relatório de Vendas - Monteiro Corretora", self.styles['CustomTitle'])
        story.append(title)
        story.append(Spacer(1, 20))
        
        # Data do relatório
        date_p = Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}", self.styles['Normal'])
        story.append(date_p)
        story.append(Spacer(1, 20))
        
        # Métricas principais
        metrics_title = Paragraph("Métricas Principais", self.styles['SectionHeader'])
        story.append(metrics_title)
        
        total_cards = len(cards)
        pipeline_stats = {}
        columns = ['atendimento_inicial', 'proposta_enviada', 'venda_andamento', 'venda_concluida', 'pos_venda']
        
        for column in columns:
            pipeline_stats[column] = len([c for c in cards if c.column == column])
        
        conversion_rate = (pipeline_stats['venda_concluida'] / max(total_cards, 1)) * 100
        
        metrics_text = f"""
        Total de operações no pipeline: {total_cards}<br/>
        Vendas concluídas: {pipeline_stats['venda_concluida']}<br/>
        Taxa de conversão: {conversion_rate:.1f}%<br/>
        Operações em andamento: {pipeline_stats['venda_andamento']}<br/>
        """
        
        story.append(Paragraph(metrics_text, self.styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Tabela do pipeline
        pipeline_title = Paragraph("Distribuição do Pipeline", self.styles['SectionHeader'])
        story.append(pipeline_title)
        
        pipeline_data = [['Etapa', 'Quantidade', 'Percentual']]
        for column in columns:
            count = pipeline_stats[column]
            percentage = (count / max(total_cards, 1)) * 100
            column_name = column.replace('_', ' ').title()
            pipeline_data.append([column_name, str(count), f"{percentage:.1f}%"])
        
        pipeline_table = Table(pipeline_data, colWidths=[3*inch, 1.5*inch, 1.5*inch])
        pipeline_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#198754')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        
        story.append(pipeline_table)
        story.append(Spacer(1, 20))
        
        # Performance mensal (se fornecida)
        if monthly_data:
            monthly_title = Paragraph("Performance Mensal", self.styles['SectionHeader'])
            story.append(monthly_title)
            
            monthly_table_data = [['Mês', 'Vendas', 'Receita', 'Ticket Médio']]
            for month in monthly_data:
                ticket_medio = month['revenue'] / max(month['sales'], 1)
                monthly_table_data.append([
                    month['month'],
                    str(month['sales']),
                    f"R$ {month['revenue']:,.2f}",
                    f"R$ {ticket_medio:,.2f}"
                ])
            
            monthly_table = Table(monthly_table_data, colWidths=[2*inch, 1.5*inch, 2*inch, 1.5*inch])
            monthly_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0dcaf0')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            
            story.append(monthly_table)
        
        # Construir PDF
        doc.build(story)
        return filepath
    
    def generate_social_media_report_pdf(self, social_data, filename=None):
        """Gerar relatório de redes sociais em PDF"""
        if not filename:
            filename = f"relatorio_redes_sociais_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        filepath = f"reports/{filename}"
        os.makedirs("reports", exist_ok=True)
        
        doc = SimpleDocTemplate(filepath, pagesize=A4)
        story = []
        
        # Título
        title = Paragraph("Relatório de Redes Sociais - Monteiro Corretora", self.styles['CustomTitle'])
        story.append(title)
        story.append(Spacer(1, 20))
        
        # Data do relatório
        date_p = Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}", self.styles['Normal'])
        story.append(date_p)
        story.append(Spacer(1, 20))
        
        # Instagram
        if 'instagram' in social_data and social_data['instagram']:
            instagram_title = Paragraph("Instagram Business", self.styles['SectionHeader'])
            story.append(instagram_title)
            
            instagram_text = "Métricas do Instagram incluindo alcance, impressões e engajamento."
            story.append(Paragraph(instagram_text, self.styles['Normal']))
            story.append(Spacer(1, 10))
        
        # Facebook
        if 'facebook' in social_data and social_data['facebook']:
            facebook_title = Paragraph("Facebook Pages", self.styles['SectionHeader'])
            story.append(facebook_title)
            
            facebook_text = "Estatísticas das páginas do Facebook conectadas."
            story.append(Paragraph(facebook_text, self.styles['Normal']))
            story.append(Spacer(1, 10))
        
        # WhatsApp
        whatsapp_title = Paragraph("WhatsApp Business", self.styles['SectionHeader'])
        story.append(whatsapp_title)
        
        whatsapp_text = "Status da integração com WhatsApp Business API."
        story.append(Paragraph(whatsapp_text, self.styles['Normal']))
        
        # Construir PDF
        doc.build(story)
        return filepath
    
    def get_report_as_base64(self, filepath):
        """Converter arquivo de relatório para base64"""
        try:
            with open(filepath, 'rb') as f:
                return base64.b64encode(f.read()).decode()
        except Exception as e:
            print(f"Erro ao converter arquivo para base64: {e}")
            return None